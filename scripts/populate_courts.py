"""
populate_courts.py — Excel dan sudlar ro'yxatini Firestore 'courts' ga yuklash.
Excel: ../sud-ro'yxat.xlsx

Chaqirish (sudqollanma_bot papkasidan):
    python populate_courts.py
"""
import sys, os, glob, re
sys.stdout.reconfigure(encoding='utf-8')

# ─── Firebase ───────────────────────────────────────────────
import firebase_admin
from firebase_admin import credentials, firestore

cred_path = 'serviceAccountKey.json'
if os.path.exists(cred_path):
    cred = credentials.Certificate(cred_path)
    firebase_admin.initialize_app(cred)
else:
    firebase_admin.initialize_app(options={'projectId': 'educationapp-4780a'})

db = firestore.client()
print("Firebase ulandi.")


# ─── Yordamchi funksiya ─────────────────────────────────────
def parse_court_names(text: str, region: str, court_type: str) -> list:
    """
    Excel katagidan sud nomlarini ajratib olish.
    Format misoli:
      "Shahar: Andijon, Xonobod
       Tumanlar: Andijon, Asaka, Baliqchi..."
    """
    results = []
    lines = str(text).strip().split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Prefix va suffix aniqlash
        suffix = ""
        if line.startswith("Shahar:"):
            suffix = "shahar sudi"
            line = line[7:].strip()
        elif line.startswith("Tumanlararo:"):
            suffix = "tumanlararo sudi"
            line = line[12:].strip()
        elif line.startswith("Tumanlar:"):
            suffix = "tuman sudi"
            line = line[9:].strip()
        elif line.startswith("Tuman:"):
            suffix = "tuman sudi"
            line = line[6:].strip()
        else:
            suffix = "sudi"

        # Vergul bilan ajratilgan nomlar
        for name in line.split(','):
            name = name.strip()
            if not name:
                continue
            full_name = f"{name} {suffix}".strip()
            results.append({
                "courtName": full_name,
                "courtType": court_type,
                "region": region,
                "isActive": True,
            })

    return results


# ─── Excel o'qish ───────────────────────────────────────────
import openpyxl

# Excel faylni topish
search_paths = ['../sud-royxat.xlsx', '../../bot-bmi/sud-royxat.xlsx']
excel_files = glob.glob('../*.xlsx') + glob.glob('../../bot-bmi/*.xlsx')

if not excel_files:
    print("XATO: Excel fayl topilmadi! ../sud-ro'yxat.xlsx mavjudligini tekshiring.")
    sys.exit(1)

excel_path = excel_files[0]
print(f"Excel fayl: {excel_path}")

wb = openpyxl.load_workbook(excel_path)
ws = wb.active
print(f"Sheet: '{wb.sheetnames[0]}', Qatorlar: {ws.max_row}")

# ─── Ma'lumotlarni tahlil qilish ────────────────────────────
courts_to_upload = []
order = 1

for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
    if row_idx == 0:
        continue  # Header

    region_raw   = row[0]
    viloyat_sud  = row[1]
    jib_text     = row[2]
    fib_text     = row[3]
    iqtisodiy    = row[4]

    if not region_raw:
        continue

    # Hudud nomini tozalash: "Andijon viloyati 7, 8" → "Andijon viloyati"
    region = re.sub(r'[\d,]+\s*$', '', str(region_raw)).strip()
    region = re.sub(r'\s+', ' ', region).strip()

    print(f"\n>>> {region}")

    # 1. Viloyat / Oliy / Shahar umumiy sudi
    if viloyat_sud:
        name = str(viloyat_sud).strip()
        courts_to_upload.append({
            "courtName": name,
            "courtType": "viloyat",
            "region": region,
            "isActive": True,
            "order": order,
        })
        print(f"  [viloyat] {name}")
        order += 1

    # 2. JIB sudlari
    if jib_text:
        jib_courts = parse_court_names(jib_text, region, "jib")
        for c in jib_courts:
            c["order"] = order
            courts_to_upload.append(c)
            print(f"  [jib] {c['courtName']}")
            order += 1

    # 3. FIB sudlari
    if fib_text:
        fib_courts = parse_court_names(fib_text, region, "fib")
        for c in fib_courts:
            c["order"] = order
            courts_to_upload.append(c)
            print(f"  [fib] {c['courtName']}")
            order += 1

    # 4. Iqtisodiy sud
    if iqtisodiy:
        name = str(iqtisodiy).strip()
        courts_to_upload.append({
            "courtName": name,
            "courtType": "iqtisodiy",
            "region": region,
            "isActive": True,
            "order": order,
        })
        print(f"  [iqtisodiy] {name}")
        order += 1


# ─── Firestorega yuklash ────────────────────────────────────
print(f"\n{'='*55}")
print(f"Jami: {len(courts_to_upload)} ta sud topildi.")

# Mavjud courts ni tozalash
print("Eski yozuvlar o'chirilmoqda...")
existing = list(db.collection('courts').stream())
batch = db.batch()
for i, doc in enumerate(existing):
    batch.delete(doc.reference)
    if (i + 1) % 400 == 0:
        batch.commit()
        batch = db.batch()
batch.commit()
print(f"  {len(existing)} ta eski yozuv o'chirildi.")

# Yangi yozuvlar yuklash
print("Yangi yozuvlar yuklanmoqda...")
batch = db.batch()
count = 0
for court in courts_to_upload:
    doc_ref = db.collection('courts').document()
    batch.set(doc_ref, court)
    count += 1
    if count % 400 == 0:
        batch.commit()
        print(f"  {count} ta yuklandi...")
        batch = db.batch()

if count % 400 != 0:
    batch.commit()

print(f"\nMuvaffaqiyat! {count} ta sud Firestore 'courts' ga yuklandi.")

# ─── Tekshiruv ──────────────────────────────────────────────
print("\n--- Namunaviy tekshiruv (Andijon viloyati) ---")
sample = db.collection('courts') \
    .where('region', '==', 'Andijon viloyati') \
    .order_by('order').limit(6).stream()
for doc in sample:
    d = doc.to_dict()
    print(f"  [{d['courtType']:10}] {d['courtName']}")

print("\n--- Region ro'yxati ---")
regions = set()
for doc in db.collection('courts').stream():
    regions.add(doc.to_dict().get('region', ''))
for r in sorted(regions):
    print(f"  {r}")
